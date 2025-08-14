from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers
from typing import Callable, Optional

from utils.logger import configurar_logger

logger = configurar_logger(__name__)

def _sanitize(s: str) -> str:
    """
    Remove caracteres inválidos para nomes de arquivo/pasta no Windows/macOS.
    """
    return "".join(c for c in str(s) if c not in '\\/:*?"<>|').strip()

class Gerador:
    """
    Gera arquivos Excel de saída a partir dos DataFrames processados por bloco.
    
    Recursos:
    - Garante a presença de colunas de data esperadas (Emissão, Vencimento, Competência).
    - Aplica formatação de data e número no Excel.
    - Cria pastas por documento e versiona arquivo se já existir.
    """

    def __init__(self, colunas_saida: list[str] | None = None):
        """
        Args:
            colunas_saida (list[str] | None): Ordem de colunas na saída.
                Se None, usa um conjunto padrão comum.
        """
        self.colunas_saida = colunas_saida or [
            "Contrato",
            "Valor",
            "Data Crédito",
            "Emissão",
            "Vencimento",
            "Competência",
        ]

    def _formatar_planilha(self, writer: pd.ExcelWriter, sheet_name: str, df_out: pd.DataFrame) -> None:
        """
        Aplica formatação simples na planilha:
        - Largura de colunas proporcional ao cabeçalho
        - Congelamento do cabeçalho
        - Formato de data e número para colunas conhecidas
        """
        ws = writer.book[sheet_name]
        ws.freeze_panes = "A2"

        # Larguras
        for i, col in enumerate(df_out.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(col) + 2))

        # Formatos
        col_idx = {c: i + 1 for i, c in enumerate(df_out.columns)}
        fmt_date = "dd/mm/yyyy"
        fmt_num = numbers.FORMAT_NUMBER_00

        for r in range(2, ws.max_row + 1):
            # Datas
            for c in ("Data Crédito", "Emissão", "Vencimento", "Competência"):
                if c in col_idx:
                    ws.cell(r, col_idx[c]).number_format = fmt_date
            # Valor com duas casas
            if "Valor" in col_idx:
                ws.cell(r, col_idx["Valor"]).number_format = fmt_num

    def gerar_arquivos_saida(self, dados: dict, pasta_destino, nome_pasta:  str | None = None, sheet_name: str = "Dados", progress_cb: Optional[Callable[[int, int, str], None]] = None,) -> list[Path]:
        """
        Gera um arquivo Excel por bloco (Documento, Plano) com colunas padronizadas.

        Args:
            dados: dict[(Documento, Plano), pd.DataFrame]
            pasta_destino: str | PathLike — pasta raiz de saída
            sheet_name: nome da planilha no Excel

        Returns:
            list[Path]: caminhos dos arquivos gerados.
        """
        out_paths: list[Path] = []
        root = Path(pasta_destino)
        root.mkdir(parents=True, exist_ok=True)

        total = len(dados)

        for idx, ((doc, plano), df) in enumerate(dados.items(), start=1):
            if df.empty:
                logger.warning(f"Sem linhas para {doc}-{plano}. Pulando geração.")
                continue

            df_out = df.copy()

            # Garante colunas de data derivadas de 'Data Crédito' se não existirem
            for col in ("Emissão", "Vencimento", "Competência"):
                if col not in df_out.columns:
                    df_out[col] = df_out.get("Data Crédito")

            # Garante todas as colunas de saída
            faltantes = [c for c in self.colunas_saida if c not in df_out.columns]
            for m in faltantes:
                df_out[m] = ""

            df_out = df_out[self.colunas_saida]

            # Pasta de destino do arquivo
            if nome_pasta:
                dir_doc = root / _sanitize(nome_pasta)
            else:
                dir_doc = root / _sanitize(doc)
            dir_doc.mkdir(parents=True, exist_ok=True)

            base = f"{_sanitize(doc)}-{_sanitize(plano)}.xlsx"
            path = dir_doc / base

            # Versiona se já existir
            k = 2
            while path.exists():
                path = dir_doc / f"{_sanitize(doc)}-{_sanitize(plano)}-v{k}.xlsx"
                k += 1

            with pd.ExcelWriter(path, engine="openpyxl", date_format="DD/MM/YYYY") as w:
                df_out.to_excel(w, index=False, sheet_name=sheet_name)
                self._formatar_planilha(w, sheet_name, df_out)

        # Progresso por arquivo
        if progress_cb:
            progress_cb(idx, total, str(path))

        logger.info(f"Arquivo gerado: {path}")
        out_paths.append(path)

        if not out_paths:
            logger.warning("Nenhum arquivo foi gerado (todos os blocos estavam vazios).")
        return out_paths
